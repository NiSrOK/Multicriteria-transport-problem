import openpyxl
from openpyxl.writer.excel import save_workbook

FILE_NAME = 'input.xlsx'
SHEET_NAME = 'Sheet'

class Table():
    wb=None
    ws=None

    def __init__(self, amount_of_suppliers, amount_of_buyers,  weight_factor_costs = None,
                 weight_factor_risks = None, weather = [], number_accidents = [],
                 road_quality = [], costs = [], needs = [], reserves = []):
        self.amount_of_suppliers = amount_of_suppliers
        self.amount_of_buyers = amount_of_buyers
        self.weight_factor_costs = weight_factor_costs
        self.weight_factor_risks = weight_factor_risks
        self.costs = costs
        self.needs = needs
        self.reserves = reserves
        self.weather = weather
        self.number_accidents = number_accidents
        self.road_quality = road_quality

    def createTable(self):
        try:
            wb = openpyxl.load_workbook(FILE_NAME)
        except:
            wb = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            wb.remove(sheet)

        # Создание нового листа
        ws = wb.create_sheet(SHEET_NAME)

        self.ws = ws
        self.wb = wb

    def updateTable(self):
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb[SHEET_NAME]

        self.ws = ws
        self.wb = wb

    def fillTable(self):
        self.ws.cell(row=1, column=1).value = f"Заполните таблицу для стоимости перевозки единицы товара между пунктами."
        # добавляем производителей
        for i in range(1, self.amount_of_suppliers + 1):
            self.ws.cell(row = i + 2, column = 1).value = f"A{i}"
        self.ws.cell(row = self.amount_of_suppliers + 3, column=1).value = f"Потребности"

        # добавляем покупателей
        for i in range(1, self.amount_of_buyers + 1):
            self.ws.cell(row = 2, column = i + 1).value = f"B{i}"
        self.ws.cell(row = 2, column=self.amount_of_buyers + 2).value = f"Запасы"

        # добавляем множители
        self.ws.cell(row = self.amount_of_suppliers + 5, column=1).value = 'Заполните весовые множители для затрат на перевозку и рисков:'
        for i in range(1, 3):
            self.ws.cell(row=self.amount_of_suppliers + 6, column=i).value = f"W{i}"
        
        # добавляем таблицу для качества дороги
        self.ws.cell(row=self.amount_of_suppliers + 9, column=1).value = f"Заполните таблицу для качества дороги между пунктами."
        self.ws.cell(row=self.amount_of_suppliers + 10, column=1).value = f"Возможные варианты : (Отличное, Хорошее, Нормальное, Плохое, Ужасное)"
        for i in range(1, self.amount_of_suppliers + 1):
            self.ws.cell(row = i + self.amount_of_suppliers + 11, column = 1).value = f"A{i}"
        for i in range(1, self.amount_of_buyers + 1):
            self.ws.cell(row=self.amount_of_suppliers + 11, column=i + 1).value = f"B{i}"
        
        # добавляем таблицу для погодных условий между пунктами
        self.ws.cell(row=self.amount_of_suppliers*2 + 13, column=1).value = f"Заполните таблицу для погодных условий между пунктами."
        self.ws.cell(row=self.amount_of_suppliers*2 + 14, column=1).value = f"Возможные варианты : (Отличное, Хорошее, Плохое)"
        for i in range(1, self.amount_of_suppliers + 1):
            self.ws.cell(row = i + self.amount_of_suppliers*2 + 15, column = 1).value = f"A{i}"
        for i in range(1, self.amount_of_buyers + 1):
            self.ws.cell(row=self.amount_of_suppliers*2 + 15, column=i + 1).value = f"B{i}"

        # добавляем таблицу аварий на участке между пунктами в год
        self.ws.cell(row=self.amount_of_suppliers*3 + 17, column=1).value = f"Заполните таблицу для аварий на участке между пунктами в год."
        self.ws.cell(row=self.amount_of_suppliers*3 + 18, column=1).value = f"Возможные варианты : (Мало, Немного, Много, Очень много)"
        for i in range(1, self.amount_of_suppliers + 1):
            self.ws.cell(row = i + self.amount_of_suppliers*3 + 19, column = 1).value = f"A{i}"
        for i in range(1, self.amount_of_buyers + 1):
            self.ws.cell(row=self.amount_of_suppliers*3 + 19, column=i + 1).value = f"B{i}"

        # Сохраняем данные
        save_workbook(self.wb, FILE_NAME)

    def readTable(self):
        # Получаем знаяения стоимостей перевозок
        for row in range(3, self.amount_of_suppliers + 3):
            costs = []
            for col in range(2, self.amount_of_buyers + 2):
                costs.append(self.ws.cell(row=row, column=col).value)
            self.costs.append(costs)

        # Получаем значения потребностей
        for col in range(2, self.amount_of_buyers + 2):
            self.needs.append(self.ws.cell(row=self.amount_of_suppliers + 3, column=col).value)

        # Получаем значения запасов
        for row in range(3, self.amount_of_suppliers + 3):
            self.reserves.append(self.ws.cell(row=row, column=self.amount_of_buyers + 2).value)
        
        # Получаем весовой множитель для затрат
        self.weight_factor_costs = float(self.ws.cell(row=self.amount_of_suppliers + 7, column=1).value)

        # Получаем весовой множитель для рисков
        self.weight_factor_risks = float(self.ws.cell(row=self.amount_of_suppliers + 7, column=2).value)

        # Получаем значения качества дорог
        for row in range(self.amount_of_suppliers + 12, self.amount_of_suppliers*2 + 12):
            road_quality = []
            for col in range(2, self.amount_of_buyers + 2):
                road_quality.append(self.ws.cell(row=row, column=col).value)
            self.road_quality.append(road_quality)
        
        # Получаем значения погодных условий между пунктами
        for row in range(self.amount_of_suppliers*2 + 16, self.amount_of_suppliers*3 + 16):
            weather = []
            for col in range(2, self.amount_of_buyers + 2):
                weather.append(self.ws.cell(row=row, column=col).value)
            self.weather.append(weather)
        
        # Получаем значения аварий между пунктами
        for row in range(self.amount_of_suppliers*3 + 20, self.amount_of_suppliers*4 + 20):
            number_accidents = []
            for col in range(2, self.amount_of_buyers + 2):
                number_accidents.append(self.ws.cell(row=row, column=col).value)
            self.number_accidents.append(number_accidents)
